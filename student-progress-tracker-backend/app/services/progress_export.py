"""
Excel export és XLSX sablon generálás (openpyxl): előrehaladás sablon letöltés és export.

A sablonban a státusz és a teljesítés féléve szerkeszthető oszlop (sárga kitöltés), többi cella zárolt;
lapvédelem opcionális jelszóval. Exportnál a fájlnév tartalmazza a NEPTUN (uid) azonosítót.
"""

import openpyxl
import io
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text
from app.utils.translations import EXPORT_HEADER, TEMPLATE_HEADER, STATUS_MAP, CATEGORY_MAP

def generate_progress_template_xlsx(user_id: int, lang: str, db: Session, current_user: dict) -> StreamingResponse:
    """
    Szakhoz tartozó összes tárgyat tartalmazó XLSX sablon generálása.

    Paraméterek:
        user_id (int): A felhasználó azonosítója.
        lang (str): Nyelv ('hu' vagy 'en').
        db (Session): SQLAlchemy adatbázis kapcsolat.
        current_user (dict): Bejelentkezett felhasználó adatai.

    Visszatérés:
        StreamingResponse: Letölthető XLSX fájl.
    """
    try:
        role_val = current_user.get("role") or current_user.get("roles")
        def is_admin(v):
            if v is None:
                return False
            if isinstance(v, (list, tuple, set)):
                return any("admin" in str(x).lower() for x in v)
            return "admin" in str(v).lower()
        if not is_admin(role_val) and int(current_user.get("user_id", -1)) != int(user_id):
            raise HTTPException(status_code=403, detail="Nincs jogosultságod más felhasználó sablonjához.")
    except KeyError:
        raise HTTPException(status_code=403, detail="Nincs jogosultságod más felhasználó sablonjához.")

    user_row = db.execute(text("SELECT major, uid FROM users WHERE id = :uid"), {"uid": user_id}).fetchone()
    if not user_row:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Felhasználó nem található.")
    major = user_row.major
    neptun = user_row.uid

    print("neptun:", neptun, "major:", major)


    courses = db.execute(text("""
        SELECT c.id, c.course_code, c.name, c.name_en, cm.type, cm.semester, cm.credit
        FROM course_major cm
        JOIN courses c ON cm.course_id = c.id
        JOIN majors m ON cm.major_id = m.id
        WHERE m.name = :major
        ORDER BY cm.semester, cm.type, c.course_code
    """), {"major": major}).fetchall()

    progress = {
        row.course_id: row for row in db.execute(text("""
            SELECT course_id, status, completed_semester
            FROM progress
            WHERE user_id = :uid
        """), {"uid": user_id}).fetchall()
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Progress template" if lang == "en" else "Előrehaladás sablon"
    ws.append(TEMPLATE_HEADER[lang])

    for c in courses:
        prog = progress.get(c.id)
        if lang == "hu":
            category = CATEGORY_MAP["hu"].get(c.type, c.type)
            status = STATUS_MAP["hu"].get(prog.status, prog.status) if prog else ""
            course_name = c.name
        else:
            category = CATEGORY_MAP["en"].get(c.type, c.type)
            status = STATUS_MAP["en"].get(prog.status, prog.status) if prog else ""
            course_name = c.name_en if c.name_en else c.name

        ws.append([
            c.course_code,
            course_name,
            category,
            c.semester,
            c.credit,
            status,
            prog.completed_semester if prog else ""
        ])

    status_col = 6
    semester_col = 7

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx, cell in enumerate(row, start=1):
            if idx in [status_col, semester_col]:
                cell.protection = openpyxl.styles.Protection(locked=False)
                cell.fill = openpyxl.styles.PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
            else:
                cell.protection = openpyxl.styles.Protection(locked=True)
            cell.border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin', color='000000'),
                right=openpyxl.styles.Side(style='thin', color='000000'),
                top=openpyxl.styles.Side(style='thin', color='000000'),
                bottom=openpyxl.styles.Side(style='thin', color='000000')
            )

    for cell in ws[1]:
        cell.border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style='thin', color='000000'),
            right=openpyxl.styles.Side(style='thin', color='000000'),
            top=openpyxl.styles.Side(style='thin', color='000000'),
            bottom=openpyxl.styles.Side(style='thin', color='000000')
        )

    ws.protection.enable()
    ws.protection.sheet = True
    ws.protection.password = "progress"

    for col in ws.columns:
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"progress_template_{neptun}_{lang}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )

def export_progress_xlsx(user_id: int, lang: str, db: Session, current_user: dict) -> StreamingResponse:
    """
    A felhasználó előrehaladásának exportja Excel (XLSX) formátumban.
    """
    if current_user["user_id"] != user_id:
        from fastapi import HTTPException
        raise HTTPException(status_code=403, detail="Nincs jogosultságod más felhasználó adatainak letöltéséhez.")

    user_row = db.execute(text("SELECT uid FROM users WHERE id = :uid"), {"uid": user_id}).fetchone()
    if not user_row:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="Felhasználó nem található.")
    neptun = user_row.uid

    rows = db.execute(text("""
        SELECT c.course_code, c.name, c.name_en, cm.credit, p.status, p.completed_semester, p.points, cm.type
        FROM progress p
        JOIN courses c ON p.course_id = c.id
        JOIN course_major cm ON cm.course_id = c.id
        WHERE p.user_id = :uid
        ORDER BY cm.type, cm.semester, c.course_code
    """), {"uid": user_id}).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Progress" if lang == "en" else "Előrehaladás"
    ws.append(EXPORT_HEADER[lang])

    for r in rows:
        if lang == "hu":
            status = STATUS_MAP["hu"].get(r.status, r.status)
            category = CATEGORY_MAP["hu"].get(r.type, r.type)
            course_name = r.name
        else:
            status = STATUS_MAP["en"].get(r.status, r.status)
            category = CATEGORY_MAP["en"].get(r.type, r.type)
            course_name = r.name_en if r.name_en else r.name

        ws.append([
            r.course_code,
            course_name,
            r.credit,
            status,
            r.completed_semester,
            r.points,
            category
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"progress_{neptun}_{lang}.xlsx"
    print("Generating export for neptun:", neptun)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )
